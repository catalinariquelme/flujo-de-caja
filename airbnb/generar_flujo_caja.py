"""
Genera un Excel de Flujo de Caja para un Airbnb.

Layout Flujo de Caja: MESES = COLUMNAS | CONCEPTOS = FILAS
Hojas: Parámetros | Flujo de Caja Mensual | Resumen
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ── Constantes ────────────────────────────────────────────────────────────────
HORIZONTE     = 36
MESES_GRACIA  = 6   # solo para colorear; la lógica usa la celda de Parámetros

C_DARK    = "1A1A2E"
C_ACCENT  = "C0392B"   # rojo oscuro (gracia)
C_BLUE    = "154360"
C_INDIGO  = "1A237E"
C_WHITE   = "FFFFFF"
C_GRAY    = "F4F6F7"
C_YELLOW  = "FFFDE7"
C_GREEN_L = "D5F5E3"
C_RED_L   = "FADBD8"
C_BLUE_L  = "D6EAF8"

CLP  = '#,##0'
PCT  = '0.0%'
PCT3 = '0.000%'


# ── Helpers de estilo ─────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def fnt(bold=False, color=C_DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_cell(cell, bg=C_WHITE, fg=C_DARK, bold=False, size=10,
               h="right", wrap=False, num_fmt=None, italic=False):
    cell.fill  = fill(bg)
    cell.font  = fnt(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = aln(h=h, wrap=wrap)
    cell.border = border_thin()
    if num_fmt:
        cell.number_format = num_fmt


# ── Referencias a Parámetros ──────────────────────────────────────────────────
# Filas reales en hoja Parámetros (se calculan en build_parametros)
P_ROW = {
    "horizonte":  4,
    "gracia":     5,
    "adr":        6,
    "noches":     7,
    "crec_adr":   8,
    "comision":  11,
    "g_comunes": 12,
    "servicios": 13,
    "fondo":     14,
    "dividendo": 15,
    "inversion": 18,
    "tasa":      19,
}


def p(key):
    """Referencia absoluta a celda de Parámetros."""
    return f"'Parámetros'!$B${P_ROW[key]}"


# ── Columnas del Flujo de Caja ────────────────────────────────────────────────
# Mes m (1-indexed) → columna B en adelante
def mc(m):
    """Column letter for month m."""
    return get_column_letter(m + 1)      # m=1→B, m=36→AK


FIRST_COL   = mc(1)                      # "B"
LAST_COL    = mc(HORIZONTE)              # "AK"
TOTAL_COL   = get_column_letter(HORIZONTE + 2)   # "AL"
FC_SHEET    = "'Flujo de Caja Mensual'"

# Filas de la hoja FC
FC = {
    "adr":          5,
    "noches":       6,
    "ing_brutos":   7,
    "comision":     8,
    "ing_netos":    9,
    "g_comunes":   11,
    "servicios":   12,
    "fondo":       13,
    "dividendo":   14,
    "tot_egresos": 15,
    "flujo_neto":  17,
    "flujo_acum":  18,
}


# ═════════════════════════════════════════════════════════════════════════════
# HOJA 1 – PARÁMETROS
# ═════════════════════════════════════════════════════════════════════════════
def build_parametros(wb):
    ws = wb.create_sheet("Parámetros")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 4
    ws.row_dimensions[1].height = 42

    # Título
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "PARÁMETROS  –  FLUJO DE CAJA AIRBNB"
    c.fill  = fill(C_DARK)
    c.font  = fnt(bold=True, color=C_WHITE, size=14)
    c.alignment = aln()

    secciones = [
        ("HORIZONTE Y OPERACIÓN", [
            ("Horizonte (meses)",               36,        None),
            ("Meses de gracia (sin ingresos)",  6,         None),
            ("ADR promedio (CLP / noche)",       38_000,    CLP),
            ("Noches promedio por mes",          21,        None),
            ("Crecimiento anual ADR",            0.03,      PCT),
        ]),
        ("COSTOS OPERATIVOS (CLP / mes)", [
            ("Comisión administración / Airbnb", 0.18,     PCT),
            ("Gastos comunes",                   90_000,   CLP),
            ("Servicios  (luz, agua, internet)", 60_000,   CLP),
            ("Fondo de mantención",              40_000,   CLP),
            ("Dividendo / arriendo",             274_000,  CLP),
        ]),
        ("INVERSIÓN Y EVALUACIÓN", [
            ("Inversión inicial  (amoblado)",    3_500_000, CLP),
            ("Tasa de descuento anual",          0.12,      PCT),
        ]),
    ]

    row = 3
    for titulo, items in secciones:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:B{row}")
        c = ws[f"A{row}"]
        c.value = titulo
        c.fill  = fill(C_BLUE)
        c.font  = fnt(bold=True, color=C_WHITE, size=10)
        c.alignment = aln(h="left")
        row += 1

        for label, valor, fmt in items:
            ws.row_dimensions[row].height = 20
            bg = C_GRAY if row % 2 == 0 else C_WHITE

            cA = ws[f"A{row}"]
            cA.value = label
            style_cell(cA, bg=bg, h="left")

            cB = ws[f"B{row}"]
            cB.value = valor
            cB.fill  = fill(C_YELLOW)
            cB.font  = fnt(bold=True, color=C_BLUE, size=11)
            cB.alignment = aln(h="right")
            cB.border = border_thin()
            if fmt:
                cB.number_format = fmt
            row += 1

        row += 1   # espacio entre secciones

    # Leyenda
    ws.row_dimensions[row].height = 28
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = "Las celdas en amarillo son editables – el resto se recalcula automáticamente."
    c.font  = fnt(italic=True, color="888888", size=9)
    c.alignment = aln(h="left", wrap=True)


# ═════════════════════════════════════════════════════════════════════════════
# HOJA 2 – FLUJO DE CAJA MENSUAL  (meses = columnas)
# ═════════════════════════════════════════════════════════════════════════════
def build_flujo(wb):
    ws = wb.create_sheet("Flujo de Caja Mensual")
    ws.sheet_view.showGridLines = False

    # ── Anchos ──
    ws.column_dimensions["A"].width = 32
    for m in range(1, HORIZONTE + 1):
        ws.column_dimensions[mc(m)].width = 11
    ws.column_dimensions[TOTAL_COL].width = 15

    # ── Fila 1: Título ──
    last_col = TOTAL_COL
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "FLUJO DE CAJA MENSUAL  –  AIRBNB"
    c.fill  = fill(C_DARK)
    c.font  = fnt(bold=True, color=C_WHITE, size=14)
    c.alignment = aln()
    ws.row_dimensions[1].height = 38

    # ── Fila 2: "Mes N" / "TOTAL" ──
    ws.row_dimensions[2].height = 28
    c = ws["A2"]
    c.value = "Concepto"
    c.fill  = fill(C_DARK)
    c.font  = fnt(bold=True, color=C_WHITE, size=10)
    c.alignment = aln(h="left")
    c.border = border_thin()

    start_date = datetime.date(2025, 1, 1)

    for m in range(1, HORIZONTE + 1):
        col = mc(m)
        in_grace = m <= MESES_GRACIA
        hdr_bg   = C_ACCENT if in_grace else C_BLUE

        c2 = ws[f"{col}2"]
        c2.value = f"Mes {m}"
        c2.fill  = fill(hdr_bg)
        c2.font  = fnt(bold=True, color=C_WHITE, size=9)
        c2.alignment = aln()
        c2.border = border_thin()

    # Total col header
    ct = ws[f"{TOTAL_COL}2"]
    ct.value = "TOTAL / FINAL"
    ct.fill  = fill(C_INDIGO)
    ct.font  = fnt(bold=True, color=C_WHITE, size=9)
    ct.alignment = aln()
    ct.border = border_thin()

    # ── Fila 3: período (Ene 2025 …) ──
    ws.row_dimensions[3].height = 20
    c = ws["A3"]
    c.value = "Período"
    c.fill  = fill("2C3E50")
    c.font  = fnt(bold=True, color=C_WHITE, size=9)
    c.alignment = aln(h="left")
    c.border = border_thin()

    for m in range(1, HORIZONTE + 1):
        month_num = ((start_date.month + m - 2) % 12) + 1
        year_num  = start_date.year + (start_date.month + m - 2) // 12
        period    = datetime.date(year_num, month_num, 1).strftime("%b %Y")
        c3 = ws[f"{mc(m)}3"]
        c3.value = period
        c3.fill  = fill("2C3E50")
        c3.font  = fnt(color=C_WHITE, size=8)
        c3.alignment = aln()
        c3.border = border_thin()

    ct3 = ws[f"{TOTAL_COL}3"]
    ct3.value = "36 meses"
    ct3.fill  = fill("2C3E50")
    ct3.font  = fnt(color=C_WHITE, size=8)
    ct3.alignment = aln()
    ct3.border = border_thin()

    # ── Secciones (filas sin datos) ──
    section_rows = {
        4:  ("INGRESOS",               C_BLUE),
        10: ("EGRESOS FIJOS",          "7B241C"),
        16: ("RESULTADO",              "1A5276"),
    }

    # ── Filas de datos ──
    # (fila, label, es_total/bold, formato)
    data_rows = [
        (5,  "ADR  (CLP / noche)",              False, CLP),
        (6,  "Noches del mes",                   False, "0"),
        (7,  "Ingresos Brutos  (CLP)",           False, CLP),
        (8,  "( - ) Comisión admin / Airbnb",    False, CLP),
        (9,  "INGRESOS NETOS  (CLP)",            True,  CLP),
        (11, "Gastos Comunes",                    False, CLP),
        (12, "Servicios  (luz, agua, internet)", False, CLP),
        (13, "Fondo de Mantención",              False, CLP),
        (14, "Dividendo / Arriendo",             False, CLP),
        (15, "TOTAL EGRESOS",                    True,  CLP),
        (17, "FLUJO NETO MENSUAL  (CLP)",        True,  CLP),
        (18, "FLUJO ACUMULADO  (CLP)",           True,  CLP),
    ]

    # Render secciones
    for row_num, (titulo, color) in section_rows.items():
        ws.row_dimensions[row_num].height = 22
        ws.merge_cells(f"A{row_num}:{TOTAL_COL}{row_num}")
        c = ws[f"A{row_num}"]
        c.value = titulo
        c.fill  = fill(color)
        c.font  = fnt(bold=True, color=C_WHITE, size=10)
        c.alignment = aln(h="left")

    # Render filas de datos
    INCOME_ROWS  = {5, 6, 7, 8, 9}
    RESULT_ROWS  = {9, 15, 17, 18}

    for row_num, label, is_bold, num_fmt in data_rows:
        ws.row_dimensions[row_num].height = 20

        # Etiqueta (col A)
        cA = ws[f"A{row_num}"]
        cA.value = label
        lbl_bg = C_BLUE_L if is_bold else C_WHITE
        style_cell(cA, bg=lbl_bg, bold=is_bold, h="left", num_fmt=None)

        # Columnas de meses
        for m in range(1, HORIZONTE + 1):
            col = mc(m)
            cell = ws[f"{col}{row_num}"]
            in_grace = m <= MESES_GRACIA

            # Color de fondo
            if in_grace and row_num in INCOME_ROWS:
                bg = C_RED_L
            elif row_num in RESULT_ROWS:
                bg = C_GREEN_L if not in_grace else C_RED_L
            elif row_num % 2 == 0:
                bg = C_GRAY
            else:
                bg = C_WHITE

            # Color de fuente
            if row_num == 18:
                fg = "1A5276"
            elif row_num == 17:
                fg = "145A32" if not in_grace else "7B241C"
            elif row_num == 9:
                fg = "145A32" if not in_grace else "7B241C"
            else:
                fg = C_DARK

            style_cell(cell, bg=bg, fg=fg, bold=is_bold, size=9, num_fmt=num_fmt)

            # Formulas
            year_idx = (m - 1) // 12

            if row_num == 5:    # ADR con crecimiento
                cell.value = (f"=IF({m}<={p('gracia')},0,"
                              f"{p('adr')}*(1+{p('crec_adr')})^{year_idx})")
            elif row_num == 6:  # Noches
                cell.value = f"=IF({m}<={p('gracia')},0,{p('noches')})"
            elif row_num == 7:  # Ingresos brutos
                cell.value = f"={col}5*{col}6"
            elif row_num == 8:  # Comisión
                cell.value = f"={col}7*{p('comision')}"
            elif row_num == 9:  # Ingresos netos
                cell.value = f"={col}7-{col}8"
            elif row_num == 11: # Gastos comunes
                cell.value = f"={p('g_comunes')}"
            elif row_num == 12: # Servicios
                cell.value = f"={p('servicios')}"
            elif row_num == 13: # Fondo
                cell.value = f"={p('fondo')}"
            elif row_num == 14: # Dividendo (0 durante gracia)
                cell.value = f"=IF({m}<={p('gracia')},0,{p('dividendo')})"
            elif row_num == 15: # Total egresos
                cell.value = f"=SUM({col}11:{col}14)"
            elif row_num == 17: # Flujo neto
                cell.value = f"={col}9-{col}15"
            elif row_num == 18: # Flujo acumulado
                if m == 1:
                    cell.value = f"=-{p('inversion')}+{col}17"
                else:
                    cell.value = f"={mc(m-1)}18+{col}17"

        # Columna TOTAL
        tc = ws[f"{TOTAL_COL}{row_num}"]
        style_cell(tc, bg=C_BLUE_L, bold=True, num_fmt=num_fmt)
        if row_num == 5:           # ADR promedio
            tc.value = f"=AVERAGE({FIRST_COL}5:{LAST_COL}5)"
        elif row_num == 18:        # Acumulado: último valor
            tc.value = f"={LAST_COL}18"
        else:                      # Suma de todos los meses
            tc.value = f"=SUM({FIRST_COL}{row_num}:{LAST_COL}{row_num})"

    # Inmovilizar: columna A + filas 1-3
    ws.freeze_panes = "B4"


# ═════════════════════════════════════════════════════════════════════════════
# HOJA 3 – RESUMEN EJECUTIVO
# ═════════════════════════════════════════════════════════════════════════════
def build_resumen(wb):
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 4

    # Título
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "RESUMEN EJECUTIVO  –  FLUJO DE CAJA AIRBNB"
    c.fill  = fill(C_DARK)
    c.font  = fnt(bold=True, color=C_WHITE, size=14)
    c.alignment = aln()
    ws.row_dimensions[1].height = 40

    tasa_mensual = f"(1+{p('tasa')})^(1/12)-1"

    # ── Rangos del FC ──
    fn_range   = f"{FC_SHEET}!{FIRST_COL}17:{LAST_COL}17"  # Flujo neto 36 meses
    acum_final = f"{FC_SHEET}!{LAST_COL}18"                 # Último acumulado

    # IRR helper: fila 50 col A (t=0) + cols B..AK (t=1..36)
    irr_range  = f"A50:{get_column_letter(HORIZONTE + 1)}50"   # A50:AK50

    kpis = [
        ("PARÁMETROS CLAVE", [
            ("Inversión inicial (amoblado)",          f"={p('inversion')}",   CLP),
            ("Horizonte de análisis",                 f"={p('horizonte')}",   "0\" meses\""),
            ("Meses de gracia (sin ingresos)",        f"={p('gracia')}",      "0"),
            ("ADR inicial (CLP / noche)",             f"={p('adr')}",         CLP),
            ("Noches promedio / mes",                 f"={p('noches')}",      "0"),
            ("Crecimiento anual ADR",                 f"={p('crec_adr')}",    PCT),
            ("Comisión admin / Airbnb",               f"={p('comision')}",    PCT),
        ]),
        ("INGRESOS ESPERADOS (mensual, sin gracia)", [
            ("Ingreso bruto mensual",
             f"={p('adr')}*{p('noches')}", CLP),
            ("Comisión mensual",
             f"={p('adr')}*{p('noches')}*{p('comision')}", CLP),
            ("Ingreso neto mensual",
             f"={p('adr')}*{p('noches')}*(1-{p('comision')})", CLP),
        ]),
        ("EGRESOS FIJOS MENSUALES", [
            ("Gastos comunes",                f"={p('g_comunes')}",  CLP),
            ("Servicios (luz, agua, internet)",f"={p('servicios')}", CLP),
            ("Fondo de mantención",           f"={p('fondo')}",      CLP),
            ("Dividendo / arriendo",          f"={p('dividendo')}",   CLP),
            ("Total egresos fijos / mes",
             f"={p('g_comunes')}+{p('servicios')}+{p('fondo')}+{p('dividendo')}",
             CLP),
            ("Margen neto mensual esperado",
             f"={p('adr')}*{p('noches')}*(1-{p('comision')})"
             f"-({p('g_comunes')}+{p('servicios')}+{p('fondo')}+{p('dividendo')})",
             CLP),
        ]),
        ("TOTALES  (36 meses)", [
            ("Ingresos netos totales",    f"={FC_SHEET}!{TOTAL_COL}9",  CLP),
            ("Total egresos",             f"={FC_SHEET}!{TOTAL_COL}15", CLP),
            ("Flujo neto total",          f"={FC_SHEET}!{TOTAL_COL}17", CLP),
            ("Flujo acumulado final",     acum_final,                   CLP),
        ]),
        ("INDICADORES FINANCIEROS", [
            ("Tasa de descuento anual",   f"={p('tasa')}",         PCT),
            ("Tasa de descuento mensual", f"={tasa_mensual}",       PCT3),
            ("VAN  (Valor Actual Neto)",
             f"=NPV({tasa_mensual},{fn_range})-{p('inversion')}",   CLP),
            ("TIR mensual",
             f"=IFERROR(IRR({irr_range}),\"N/D\")",                 "0.00%"),
            ("TIR anual equiv.",
             f"=IFERROR((1+IRR({irr_range}))^12-1,\"N/D\")",        PCT),
            ("Payback (meses aprox.)",
             f"=IFERROR(IF(COUNTIF({FC_SHEET}!{FIRST_COL}18:{LAST_COL}18,\"<0\")={HORIZONTE},"
             f"\"No recuperado en el horizonte\","
             f"COUNTIF({FC_SHEET}!{FIRST_COL}18:{LAST_COL}18,\"<0\")&\" meses\"),\"N/D\")",
             "@"),
        ]),
    ]

    row = 3
    for titulo, items in kpis:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:B{row}")
        c = ws[f"A{row}"]
        c.value = titulo
        c.fill  = fill(C_BLUE)
        c.font  = fnt(bold=True, color=C_WHITE, size=10)
        c.alignment = aln(h="left")
        row += 1

        for label, formula, num_fmt in items:
            ws.row_dimensions[row].height = 20
            bg = C_GRAY if row % 2 == 0 else C_WHITE

            cA = ws[f"A{row}"]
            cA.value = label
            style_cell(cA, bg=bg, h="left")

            cB = ws[f"B{row}"]
            cB.value = formula
            cB.fill  = fill(bg)
            cB.font  = fnt(bold=True, color=C_DARK, size=11)
            cB.alignment = aln(h="right")
            cB.border = border_thin()
            if num_fmt and num_fmt != "@":
                cB.number_format = num_fmt
            row += 1

        row += 1   # espacio entre secciones

    # ── Fila 50: helper para IRR [-Inv, FN_1 .. FN_36] ──
    # t=0 → col A; t=1..36 → cols B..AK
    ws.row_dimensions[50].height = 14
    lbl = ws["A50"]
    lbl.value = "Helper TIR →"
    lbl.font  = fnt(italic=True, color="AAAAAA", size=8)
    lbl.alignment = aln(h="left")

    ws["B50"].value = f"=-{p('inversion')}"          # t = 0
    for m in range(1, HORIZONTE + 1):
        col_helper = get_column_letter(m + 2)         # t=1→C, t=36→AL
        fc_m_col   = mc(m)                            # FC sheet col for month m
        ws[f"{col_helper}50"].value = f"={FC_SHEET}!{fc_m_col}17"

    # Actualizar irr_range para incluir t=0 en col B y t=36 en col AL
    # B50 = t=0 (-inv), C50..AL50 = t=1..t=36  → range B50:AL50 (37 celdas)
    irr_range_real = f"B50:{get_column_letter(HORIZONTE + 2)}50"   # B50:AL50

    # Corregir las fórmulas TIR que usan irr_range
    # Buscar y actualizar las celdas de TIR en la hoja
    for r in range(1, 50):
        for col_idx in [2]:  # col B
            cell = ws.cell(row=r, column=col_idx)
            if cell.value and isinstance(cell.value, str) and irr_range in cell.value:
                cell.value = cell.value.replace(irr_range, irr_range_real)

    # Nota al pie
    row += 2
    ws.row_dimensions[row].height = 36
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = ("Nota: VAN > 0 indica que el proyecto supera la rentabilidad mínima exigida. "
               "TIR se calcula con la inversión inicial en t=0 y flujos netos mensuales t=1..36.")
    c.font  = fnt(italic=True, color="888888", size=9)
    c.alignment = aln(h="left", wrap=True)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_parametros(wb)
    build_flujo(wb)
    build_resumen(wb)

    out = "/home/user/flujo-de-caja/airbnb/flujo_caja_airbnb.xlsx"
    wb.save(out)
    print(f"Archivo guardado: {out}")


if __name__ == "__main__":
    main()
